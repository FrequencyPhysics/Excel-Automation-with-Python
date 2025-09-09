import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_workbook(input_file, output_file):
    # Load workbook and select sheet
    wb = xl.load_workbook(input_file)
    sheet = wb['Sheet1']

    print(f"Opened file: {input_file}")
    print(f"Sheet'{sheet.title}' has {sheet.max_row} rows")

    # Loops through rows to apply price correction
    for row in range(2,sheet.max_row + 1): # Skip header row
        price_cell = sheet.cell(row,3) # Price column - 3
        corrected_price = price_cell.value * 0.9 # Apply 10% Discount
        sheet.cell(row,4).value = corrected_price # Price corrected column - 4

    # Creating Bar Chart to visualize corrected prices
    values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4)
    chart = BarChart()
    chart.title = "Corrected Prices"
    chart.x_axis.title = "Items"
    chart.y_axis.title = "Price"
    chart.add_data(values, titles_from_data=False)
    sheet.add_chart(chart, 'e2')

    wb.save(output_file)
    print(f"Processed file saved as: {output_file}")

if __name__ == "__main__":
    # Default file names
    input_file = "input.xlsx"
    output_file = "output.xlsx"

    process_workbook(input_file, output_file)